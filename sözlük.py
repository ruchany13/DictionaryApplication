from openpyxl import load_workbook
import random
from datetime import datetime
import time




wb = load_workbook("sözlük.xlsx")
ws = wb.active

#TODO yanlýþ yazdýðým sorularý silme yap
def read_all_excel(column):
    liste = []
    sözlük = {}
    # 2. sütündaki kelimeleri liste içine atıyoruz.
    for i in range(2, ws.max_row + 1): #row(2 - max) i = 2, 3, 4, 5, 6 ...
        liste.append(ws.cell(i, column).value)
        sözlük[ws.cell(i, column).value] = ws.cell(i, 1).value
    return liste, sözlük


# TODO Eklenme zamanıda eklencek. Time modülü ile eklenen gün.

# Kullanıcının eklemek istediği kelimenin sistemde daha önceden olup olmadığını kontrol ediyoruz.
def kelime_kayıt_kontrol(kelime, word_type):
    liste, sözlük = read_all_excel(2)

    # TODO İf the same word have two type ( drill : noun ,verb) , you change system. İf word and type same...
    # Sözlüğe kelimenin numarasını eşleştiriyoruz. Bu sayede kayıt satırına ekleme yapabiliriz.(+1)


    if kelime in liste:
        word_number = sözlük[kelime]
        print("Word number:", word_number)
        print("Dictionary", sözlük)
        variable = ws.cell(word_number + 1, 3).value  # Word's that before save type'
        if word_type == variable:
            kelime_numarası = sözlük[kelime]
            # Kelimenin sıra numrasını okunuyor. Bir eklenip konumu bulunuyor. Konumundaki numara okunup tekrar 1 fazla ile yazılır.
            kelime_konumu = "H" + str(kelime_numarası + 1)
            numara = ws.cell(kelime_numarası + 1, 8).value
            numara += 1
            ws[kelime_konumu] = numara
            wb.save("sözlük.xlsx")
            return 0
        else:
            return 1

    else:
        return 1



def word_chooser():
    word_list, word_dict = read_all_excel(2)
    a = ws.cell(ws.max_row, 1).value  # Found last word's number
    # a = a/2
    chosen_number = random.randint(1, a)  # TODO + random.randint(1, a)  # Create random number
    chosen_word = word_list[chosen_number - 1]  # Chose random word in the excel second column
    # TODO add a number in the sorulan
    chosen_word_number = word_dict[chosen_word] + 1  # You have to add 1 due to first column have information
    asking_number = int(ws.cell(chosen_word_number, 7).value)  # How many times program ask chosen_word ?
    asking_number += 1  # Program ask one more.
    chosen_word_location = "G" + str(chosen_word_number)  # Write new asking value,
    chosen_word_translate = ws.cell(chosen_word_number, 6).value
    ws[chosen_word_location] = asking_number
    wb.save("sözlük.xlsx")
    
    return chosen_word, asking_number, chosen_word_translate


def kelime_ekleme(kelime, türü, eş=None, zıt=None, anlam=None, sentence=None):
    if kelime_kayıt_kontrol(kelime, türü) == 1:
        bugun = datetime.today()
        tarih = datetime.strftime(bugun, '%x')
        ws.append([ws.max_row, kelime, türü, eş, zıt, anlam, 0, 1, tarih, sentence])
        wb.save("sözlük.xlsx")


def kelime_sorgulama():
    kelime = input("Kelimenin İngilizcesini giriniz:")
    print("-" * 75)
    kelimenin_türü = input("Kelimenin türünü giriniz:")
    print("-" * 75)
    kontrol = kelime_kayıt_kontrol(kelime, kelimenin_türü)
    if kontrol == 0:
        print("Maalesef kelime sözlükte bulunmaktadır. Başka bir kelime deneyiniz!!!")
    elif kontrol == 1:
        anlam = input("Kelimenin Türkçe anlamını giriniz:")
        print("-" * 75)
        eş_anlamlısı = input("Varsa kelimenizin eş anlamlısını giriniz:")
        print("-" * 75)
        zıt_anlamlısı = input("Varsa kelimenin zıt anlamlısını giriniz:")
        print("-" * 75)
        sentence = input("Örnek bir cümle giriniz: ")
        kelime_ekleme(kelime, kelimenin_türü, anlam=anlam, sentence=sentence)
        print("*" * 75)


def game():
    pass



print("#" * 66, "[-][o][x]")
print("Benim Sözlüğüm Uygulamasına Hoşgeldiniz!! Ruchan Yalçın 2021")
program = ("""1 for add word
2 for exercise
'q' for exit  :""")

program = input(program)

i = 0
while True:
    if program == "q":
        break
    elif program == "1":
        kelime_sorgulama()
    elif program == "2":
        # TODO add : user's thinking time && ask a question that "Can you remember word?"
        # TODO add : If user remember word click 1/yes nor click 2/no
        # TODO maybe : If user remember word , add 1 point for user. İf user not , remove 1 point.
        chosen_word, asking_number, translate = word_chooser()
        print(chosen_word)
        answer = input("Do you remember ? ")
        if answer == "":
            print("Wov!Check it.")
            
        time.sleep(5)
        # TODO write the word's translate
        print(translate )
        time.sleep(0.5)
        print("{} is asked {} times.".format(chosen_word, asking_number))
        i += 1
